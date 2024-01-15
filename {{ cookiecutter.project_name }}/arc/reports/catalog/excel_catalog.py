# -*- coding: utf-8 -*-
"""
Module for the generation of steps catalogue in Excel format.
"""
import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from arc.contrib.steps.host import host_keywords
from arc.settings.settings_manager import Settings
from arc.contrib.steps.api import api_keywords
from arc.contrib.steps.general import functional_keywords, data_keywords, ftp_keywords, mail_keywords
from arc.contrib.steps.web import web_keywords, appian_keywords
from arc.reports.catalog.pydoc_formatter import get_pydoc_info
from arc.reports.catalog.user_step import get_list_user_steps

logger = logging.getLogger(__name__)

PATH_HOME = Settings.OUTPUT_PATH.get() + os.sep


class Catalog:
    """
    Control class that generates an Excel catalogue of default steps and user steps.
    """
    workbook = None
    functions_list = []
    options_talos = {}
    all_default_function_config = {}
    user_functions_list = []
    all_user_function_config = {}
    step_verbs = ["step", "given", "when", "then", "and"]

    def __init__(self):
        self.workbook = Workbook()
        self.get_config()
        self.get_user_function_data()
        self.write_function_user_data_in_excel()
        self.get_default_function_data()
        self.write_function_default_data_in_excel()
        self.word_finder()
        self.save_excel()

    def set_sheet_title(self, title_name):
        """
        Creates the title of the Excel sheets according to the name passed as a parameter.
        :param title_name:
        :return:
        """
        logger.debug(f'Creating the title of the steps catalogue sheet: {title_name}')
        worksheet = self.workbook[title_name]
        worksheet['A1'] = title_name
        worksheet.merge_cells('A1:F1')
        font = Font(color="FF0000", bold=True, sz=18)
        alignment = Alignment(horizontal="center", vertical="center")
        a1 = worksheet['A1']
        a1.font = font
        a1.alignment = alignment

    def set_headers_title(self, sheet):
        """
        Creates the default headers contained in the Excel with the information of the step columns.
        :param sheet:
        :return:
        """
        logger.debug(f'Creating the headers of the steps catalogue sheet: {sheet}')
        worksheet = self.workbook[sheet]
        font = Font(bold=True, sz=12, color="FF0000")
        alignment = Alignment(horizontal="center", vertical="center")
        worksheet['A1'] = "Type"
        worksheet['A1'].font = font
        worksheet['A1'].alignment = alignment
        worksheet['B1'] = "Step"
        worksheet['B1'].font = font
        worksheet['B1'].alignment = alignment
        worksheet['C1'] = "Step Name"
        worksheet['C1'].font = font
        worksheet['C1'].alignment = alignment
        worksheet['D1'] = "Information"
        worksheet['D1'].font = font
        worksheet['D1'].alignment = alignment
        worksheet['E1'] = "Params"
        worksheet['E1'].font = font
        worksheet['E1'].alignment = alignment
        worksheet['F1'] = "Step Example"
        worksheet['F1'].font = font
        worksheet['F1'].alignment = alignment
        worksheet['G1'] = "File Path"
        worksheet['G1'].font = font
        worksheet['G1'].alignment = alignment

    @staticmethod
    def set_column_width(worksheet):
        """
        Set column width depending on the dimensions fo cells.
        :param worksheet:
        :return:
        """
        ws = worksheet
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value + 2

    @staticmethod
    def set_column_custom_width(worksheet, column, width):
        """
        Set column width custom.
        :param worksheet:
        :param column:
        :param width:
        :return:
        """
        worksheet.column_dimensions[column].width = width

    def get_user_function_data(self):
        """
        Gets all steps functions from the user.
        :return:
        """
        if self.options_talos["user_steps"]:
            logger.debug("Obtaining all user steps")
            self.user_functions_list = get_list_user_steps()

            titles = []
            for _list in self.user_functions_list:
                for _dict in _list:
                    if _dict["Function Name"] not in titles:
                        titles.append(_dict["Function Name"])

                for title in titles:
                    if "'Function Name': '" + title + "'" in str(_list):
                        self.all_user_function_config[title] = {"Title": title, "Data": _list}

    def write_function_user_data_in_excel(self):
        """
        Write into Excel catalogue user data functions.
        :return:
        """
        logger.debug("Writing in the steps catalogue the data of the user functions")
        font = Font(bold=True, sz=12)
        for function_key in self.all_user_function_config.keys():
            title = ""
            all_functions = []
            for each_key in self.all_user_function_config[function_key].keys():
                if each_key == "Title":
                    title = self.all_user_function_config[function_key][each_key]

                if each_key == "Data":
                    all_functions = self.all_user_function_config[function_key][each_key]

                if "steps" not in str(title).lower() or "step" not in str(title).lower():
                    title = title + " Steps"
                self.create_sheet(title)

            worksheet = self.workbook[title]  # noqa
            self.set_headers_title(title)
            cont = 2
            for list_with_dict in all_functions:
                function_tag = list_with_dict["Tags"]
                verb_step = list_with_dict["Verb Step"]
                step = list_with_dict["Step"]
                py_doc = list_with_dict["Py Doc"]
                example = list_with_dict["Example"]
                params = list_with_dict["Params"]
                file_paths = list_with_dict["Function Path"]
                if str(verb_step).lower() in self.step_verbs:
                    if function_tag is None:
                        function_tag = "None"
                    worksheet['A' + str(cont)] = function_tag
                    worksheet['A' + str(cont)].font = font
                    worksheet['A' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    worksheet['B' + str(cont)] = verb_step
                    worksheet['B' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    worksheet['C' + str(cont)] = str(step).replace("(u'", "").replace("')", "").replace("(\"", "")
                    worksheet['C' + str(cont)].alignment = Alignment(vertical="center")

                    worksheet['D' + str(cont)].alignment = Alignment(vertical="center", wrapText=True)
                    text_doc = ""
                    for doc in py_doc:
                        text_doc = text_doc + doc + "\n"
                    worksheet['D' + str(cont)] = text_doc[:-1]

                    worksheet['E' + str(cont)].alignment = Alignment(vertical="center", horizontal="center",
                                                                     wrapText=True)
                    text_param = ""
                    for param in params:
                        text_param = text_param + param + "\n"
                    worksheet['E' + str(cont)] = text_param[:-1]

                    example_list = str(example).split("\n")
                    worksheet['F' + str(cont)].alignment = Alignment(vertical="center", horizontal="center",
                                                                     wrapText=True)
                    text_example = ""
                    for example_lines in example_list:
                        text_example = text_example + example_lines + "\n"
                    worksheet['F' + str(cont)] = text_example[:-1]

                    font_file = Font(bold=False, sz=12)
                    worksheet['G' + str(cont)] = file_paths
                    worksheet['G' + str(cont)].font = font_file
                    worksheet['G' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    cont += 1

            self.set_column_width(worksheet)
            self.set_column_custom_width(worksheet, "D", 40)
            self.set_column_custom_width(worksheet, "F", 40)

    def get_default_function_data(self):
        """
        Save into all default function config dict the options about Talos default steps.
        :return:
        """
        logger.debug("Getting Talos default steps data")
        if self.options_talos["default_api"]:
            logger.debug("Getting Talos default api steps data")
            self.all_default_function_config["default_api"] = {
                "Title": "Api Default Steps",
                "Data": get_pydoc_info(api_keywords)
            }
        if self.options_talos["default_web"]:
            logger.debug("Getting Talos default web steps data")
            self.all_default_function_config["default_web"] = {
                "Title": "Web Default Steps",
                "Data": get_pydoc_info(web_keywords)
            }

        if self.options_talos["default_functional"]:
            logger.debug("Getting Talos default functional steps data")
            self.all_default_function_config["functional_keywords"] = {
                "Title": "Functional Default Steps",
                "Data": get_pydoc_info(functional_keywords)
            }
        if self.options_talos["default_data"]:
            logger.debug("Getting Talos default data steps data")
            self.all_default_function_config["default_data"] = {
                "Title": "Data Default Steps",
                "Data": get_pydoc_info(data_keywords)
            }

        if self.options_talos["default_ftp"]:
            logger.debug("Getting Talos default ftp steps data")
            self.all_default_function_config["default_ftp"] = {
                "Title": "FTP Default Steps",
                "Data": get_pydoc_info(ftp_keywords)
            }

        if self.options_talos["default_appian"]:
            logger.debug("Getting Talos default appian steps data")
            self.all_default_function_config["default_appian"] = {
                "Title": "Appian Default Steps",
                "Data": get_pydoc_info(appian_keywords)
            }

        if self.options_talos["default_host"]:
            logger.debug("Getting Talos default host steps data")
            self.all_default_function_config["default_host"] = {
                "Title": "Host Default Steps",
                "Data": get_pydoc_info(host_keywords)
            }

        if self.options_talos["default_mail"]:
            logger.debug("Getting Talos default mail steps data")
            self.all_default_function_config["default_mail"] = {
                "Title": "Mail Default Steps",
                "Data": get_pydoc_info(mail_keywords)
            }

    def write_function_default_data_in_excel(self):
        """
        Write into catalogue default steps data.
        :return:
        """
        logger.debug("Writing in the steps catalogue Talos default steps")
        font = Font(bold=True, sz=12)
        for function_key in self.all_default_function_config.keys():
            title = ""
            all_functions = []
            for each_key in self.all_default_function_config[function_key].keys():
                if each_key == "Title":
                    title = self.all_default_function_config[function_key][each_key]
                if each_key == "Data":
                    all_functions = self.all_default_function_config[function_key][each_key]
                self.create_sheet(title)

            worksheet = self.workbook[title]  # noqa
            self.set_headers_title(title)
            cont = 2
            for list_with_dict in all_functions:
                function_tag = list_with_dict["Tags"]
                verb_step = list_with_dict["Verb Step"]
                step = list_with_dict["Step"]
                py_doc = list_with_dict["Py Doc"]
                example = list_with_dict["Example"]
                params = list_with_dict["Params"]
                file_paths = list_with_dict["Function Path"]
                if str(verb_step).lower() in self.step_verbs:
                    if function_tag is None:
                        function_tag = "None"
                    worksheet['A' + str(cont)] = function_tag
                    worksheet['A' + str(cont)].font = font
                    worksheet['A' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    worksheet['B' + str(cont)] = verb_step
                    worksheet['B' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    worksheet['C' + str(cont)] = str(step).replace("(u'", "").replace("')", "").replace("(\"", "")
                    worksheet['C' + str(cont)].alignment = Alignment(vertical="center")

                    worksheet['D' + str(cont)].alignment = Alignment(vertical="center", wrapText=True)
                    text_doc = ""
                    for doc in py_doc:
                        text_doc = text_doc + doc + "\n"
                    worksheet['D' + str(cont)] = text_doc[:-1]

                    worksheet['E' + str(cont)].alignment = Alignment(vertical="center", horizontal="center",
                                                                     wrapText=True)
                    text_param = ""
                    for param in params:
                        text_param = text_param + param + "\n"
                    worksheet['E' + str(cont)] = text_param[:-1]

                    example_list = str(example).split("\n")
                    worksheet['F' + str(cont)].alignment = Alignment(vertical="center", horizontal="center",
                                                                     wrapText=True)
                    text_example = ""
                    for example_lines in example_list:
                        text_example = text_example + example_lines + "\n"
                    worksheet['F' + str(cont)] = text_example[:-1]

                    font_file = Font(bold=False, sz=12)
                    worksheet['G' + str(cont)] = file_paths
                    worksheet['G' + str(cont)].font = font_file
                    worksheet['G' + str(cont)].alignment = Alignment(horizontal="center", vertical="center")

                    cont += 1

            self.set_column_width(worksheet)
            self.set_column_custom_width(worksheet, "D", 40)
            self.set_column_custom_width(worksheet, "F", 40)

    def create_sheet(self, sheet):
        """
        Create a new sheet.
        :param sheet:
        :return:
        """
        logger.debug(f"Creating new sheet: {sheet}")
        if sheet not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet)

        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]

        self.set_headers_title(sheet)

    def get_config(self):
        """
        Save into option_talos the options configured in settings file.
        :return:
        """
        self.options_talos = {
            "default_api": Settings.PYTALOS_CATALOG.get('steps').get('default_api'),
            "default_web": Settings.PYTALOS_CATALOG.get('steps').get('default_web'),
            "default_functional": Settings.PYTALOS_CATALOG.get('steps').get('default_functional'),
            "default_data": Settings.PYTALOS_CATALOG.get('steps').get('default_data'),
            "default_ftp": Settings.PYTALOS_CATALOG.get('steps').get('default_ftp'),
            "default_appian": Settings.PYTALOS_CATALOG.get('steps').get('default_appian'),
            "default_host": Settings.PYTALOS_CATALOG.get('steps').get('default_host'),
            "default_mail": Settings.PYTALOS_CATALOG.get('steps').get('default_mail'),
            "default_autogui": Settings.PYTALOS_CATALOG.get('steps').get('default_autogui'),
            "user_steps": Settings.PYTALOS_CATALOG.get('steps').get('user_steps'),
        }

    def word_finder(self):
        """
        Find the gherkin reserved words to enter into the formatted excel.
        :return:
        """
        values = ["STEP", "GIVEN", "WHEN", "THEN", "AND", "None"]

        font_verb_step = Font(color="5C0303", bold=True, sz=10)
        font_verb_given = Font(color="030E5C", bold=True, sz=10)
        font_verb_when = Font(color="035C13", bold=True, sz=10)
        font_verb_then = Font(color="5C0353", bold=True, sz=10)
        font_verb_and = Font(color="03595C", bold=True, sz=10)
        font_verb_none = Font(color="ED0606", bold=True, sz=10)

        for sheet in self.workbook.sheetnames:
            worksheet = self.workbook[sheet]
            for i in range(2, worksheet.max_row + 1):
                for j in range(1, worksheet.max_column + 1):
                    for value in values:
                        if str(value) == str(worksheet.cell(row=i, column=j).value):
                            if value == "STEP":
                                worksheet.cell(row=i, column=j).font = font_verb_step
                            if value == "GIVEN":
                                worksheet.cell(row=i, column=j).font = font_verb_given
                            if value == "WHEN":
                                worksheet.cell(row=i, column=j).font = font_verb_when
                            if value == "THEN":
                                worksheet.cell(row=i, column=j).font = font_verb_then
                            if value == "AND":
                                worksheet.cell(row=i, column=j).font = font_verb_and
                            if value == "None":
                                worksheet.cell(row=i, column=j).font = font_verb_none

    def save_excel(self):
        """
        Save catalogue generated.
        :return:
        """
        output_path_catalog = str(PATH_HOME + Settings.PYTALOS_CATALOG.get('excel_file_name') + ".xlsx").replace("\\", "/")
        logger.info(f'Catalogue of steps generated in: {output_path_catalog}')
        self.workbook.save(filename=output_path_catalog)
